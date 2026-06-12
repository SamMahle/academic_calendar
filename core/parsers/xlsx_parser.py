"""XLSX syllabus parser using openpyxl.

Produces both stringified tables (for the generic extractor) and typed
SheetData rows (for the structured syllabus extractor, which needs real
datetime cells and sheet names).
"""

import io
from pathlib import Path
from typing import Union

from openpyxl import load_workbook

from core.parsers import ParsedDoc, SheetData


def parse_xlsx(source: Union[Path, bytes, io.BytesIO]) -> ParsedDoc:
    """Parse an XLSX file; return all non-empty sheet content."""
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    elif isinstance(source, Path):
        source = source.open("rb")

    wb = load_workbook(source, data_only=True)
    result = ParsedDoc()
    lines: list[str] = []

    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        raw_rows: list[list] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)
                raw_rows.append(list(row))
                joined = "\t".join(c for c in cells if c)
                if joined:
                    lines.append(joined)
        if rows:
            result.tables.append(rows)
            result.sheets.append(SheetData(name=sheet.title, rows=raw_rows))

    result.full_text = "\n".join(lines)
    result.paragraphs = [ln for ln in lines if ln]
    return result
