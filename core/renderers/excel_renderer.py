"""Week-grid Excel renderer.

Layout (8 columns):
  A         B–H
  Month     Mon Tue Wed Thu Fri Sat Sun

For each calendar week, six rows are written (always uniform):
  1. Date-number row  (day of month + special-day note)
  2. Day-type row     (D-1, D-2, HOL, TEE, BRK …)
  3–6. Four fixed event slots

Column A is merged across all rows that belong to the same calendar month,
and alternates between two soft colors so months are easy to distinguish.
"""

import io
import math
from datetime import date, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.base_calendar import BaseCalendar
from core.models import Course, Event

_DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
_MONTH_COL = 1      # column A
_FIRST_DAY_COL = 2  # column B (Monday)

_DAY_COL_WIDTH = 22
_N_EVENT_ROWS = 4                 # fixed slots per week
_ROWS_PER_WEEK = 2 + _N_EVENT_ROWS  # date row + type row + 4 event rows = 6
_CHARS_PER_LINE = 23              # chars that fit in a width-22 col at 9pt
_LINE_HEIGHT = 13                 # points per wrapped text line

# Alternating month-column fills — one pair, cycles by month index
_MONTH_FILLS = ["C5D9F1", "D8E4BC"]   # soft blue / soft green


def _lines_needed(text: str) -> int:
    total = 0
    for part in str(text).split("\n"):
        total += max(1, math.ceil(len(part) / _CHARS_PER_LINE))
    return total


def _thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _day_label_and_fills(meta, theme: dict) -> tuple[str, str, str]:
    """Return (label, fill_hex, font_hex) for the day-type row."""
    if meta is None:
        return "", theme["weekend_fill"], theme["weekend_font"]
    dt = meta.day_type
    if dt == "1":
        return "D-1", theme["day1_marker"], "2E75B6"
    if dt == "2":
        return "D-2", theme["day2_marker"], "375623"
    if dt == "holiday":
        return "HOL", theme["holiday_fill"], theme["holiday_font"]
    if dt == "tee":
        return "TEE", theme["tee_fill"], theme["tee_font"]
    if dt == "grad":
        return "GRAD", theme["grad_fill"], theme["grad_font"]
    if dt == "break":
        return "BRK", theme["holiday_fill"], theme["holiday_font"]
    return "", theme["weekend_fill"], theme["weekend_font"]


def render_excel(
    events: list[Event],
    courses: list[Course],
    calendar: BaseCalendar,
    theme: dict,
    ay: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = ay[:31]

    # Column widths
    ws.column_dimensions[get_column_letter(_MONTH_COL)].width = 7
    for i in range(7):
        ws.column_dimensions[get_column_letter(_FIRST_DAY_COL + i)].width = _DAY_COL_WIDTH

    # Event lookup
    by_date: dict[date, list[Event]] = {}
    for ev in events:
        by_date.setdefault(ev.date, []).append(ev)

    course_map: dict[str, Course] = {c.code: c for c in courses}

    cal_days = sorted(calendar.days())
    if not cal_days:
        raise ValueError("Calendar has no days")
    start_mon = _monday(cal_days[0])
    last_day = cal_days[-1]

    # Ordered list of week Monday dates
    weeks: list[date] = []
    w = start_mon
    while w <= last_day:
        weeks.append(w)
        w += timedelta(weeks=1)

    # ── Pre-compute month spans ─────────────────────────────────────────────
    # Title row = 1, header row = 2, content starts at row 3.
    # Each week always occupies exactly _ROWS_PER_WEEK rows.
    CONTENT_START = 3

    # month_info: ordered list of {label, first_row, last_row, color_idx}
    month_info: dict[str, dict] = {}   # keyed by "MMM" label
    month_order: list[str] = []

    for idx, week_mon in enumerate(weeks):
        first_row = CONTENT_START + idx * _ROWS_PER_WEEK
        last_row = first_row + _ROWS_PER_WEEK - 1
        label = week_mon.strftime("%b").upper()
        if label not in month_info:
            month_info[label] = {
                "first_row": first_row,
                "last_row": last_row,
                "color_idx": len(month_order),
            }
            month_order.append(label)
        else:
            month_info[label]["last_row"] = last_row

    # ── Title row ──────────────────────────────────────────────────────────
    row = 1
    tc = ws.cell(row=row, column=_MONTH_COL, value=f"Cadet Calendar — {ay}")
    tc.font = Font(bold=True, size=14, color=theme["header_font"])
    tc.fill = _fill(theme["header_fill"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Day-name header row ────────────────────────────────────────────────
    ws.cell(row=row, column=_MONTH_COL, value="").fill = _fill(theme["day_header_fill"])
    for i, name in enumerate(_DAYS):
        c = ws.cell(row=row, column=_FIRST_DAY_COL + i, value=name)
        c.font = Font(bold=True, color=theme["day_header_font"])
        c.fill = _fill(theme["day_header_fill"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    ws.row_dimensions[row].height = 18
    row += 1  # row is now CONTENT_START

    # ── Weeks ──────────────────────────────────────────────────────────────
    for week_mon in weeks:
        week = [week_mon + timedelta(days=i) for i in range(7)]

        # ── Date-number row ────────────────────────────────────────────────
        date_row = row
        date_row_lines = 1
        for i, d in enumerate(week):
            col = _FIRST_DAY_COL + i
            meta = calendar.get_day_meta(d) if d <= last_day else None
            dt = meta.day_type if meta else "weekend"

            if dt == "weekend":
                fgc, fnt = theme["weekend_fill"], theme["weekend_font"]
            elif dt in ("holiday", "break"):
                fgc, fnt = theme["holiday_fill"], theme["holiday_font"]
            elif dt == "tee":
                fgc, fnt = theme["tee_fill"], theme["tee_font"]
            elif dt == "grad":
                fgc, fnt = theme["grad_fill"], theme["grad_font"]
            elif dt == "1":
                fgc, fnt = theme["day1_marker"], "000000"
            elif dt == "2":
                fgc, fnt = theme["day2_marker"], "000000"
            else:
                fgc, fnt = "FFFFFF", "000000"

            day_val = d.day if d <= last_day else ""
            note = ""
            if meta and meta.notes:
                note = f"\n{meta.notes[0]}"

            c = ws.cell(row=date_row, column=col, value=f"{day_val}{note}" if note else day_val)
            c.font = Font(bold=True, size=10, color=fnt)
            c.fill = _fill(fgc)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=bool(note))
            c.border = _thin()
            if note:
                date_row_lines = max(date_row_lines, _lines_needed(f"{day_val}{note}"))
        ws.row_dimensions[date_row].height = max(18, date_row_lines * _LINE_HEIGHT + 4)

        # ── Day-type row ───────────────────────────────────────────────────
        type_row = row + 1
        for i, d in enumerate(week):
            col = _FIRST_DAY_COL + i
            meta = calendar.get_day_meta(d) if d <= last_day else None
            label, fgc, fnt = _day_label_and_fills(meta, theme)
            c = ws.cell(row=type_row, column=col, value=label)
            c.font = Font(size=8, bold=True, color=fnt)
            c.fill = _fill(fgc)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()
        ws.row_dimensions[type_row].height = 14

        # ── Event slots (always 4) ─────────────────────────────────────────
        for slot in range(_N_EVENT_ROWS):
            evt_row = row + 2 + slot
            slot_lines = 1
            for i, d in enumerate(week):
                col = _FIRST_DAY_COL + i
                day_evts = by_date.get(d, [])
                meta = calendar.get_day_meta(d) if d <= last_day else None
                dt = meta.day_type if meta else "weekend"

                if slot < len(day_evts):
                    ev = day_evts[slot]
                    crs = course_map.get(ev.course_code)
                    color = crs.color if crs else "CCCCCC"
                    label = f"{ev.course_code} {ev.title}"
                    c = ws.cell(row=evt_row, column=col, value=label)
                    c.fill = _fill(color)
                    c.font = Font(size=9, bold=True, color="000000")
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = _thin()
                    slot_lines = max(slot_lines, _lines_needed(label))
                else:
                    c = ws.cell(row=evt_row, column=col, value="")
                    if dt in ("weekend", "R"):
                        c.fill = _fill(theme["weekend_fill"])
                    elif dt in ("holiday", "break"):
                        c.fill = _fill(theme["holiday_fill"])
                    elif dt == "tee":
                        c.fill = _fill(theme["tee_fill"])
                    elif dt == "grad":
                        c.fill = _fill(theme["grad_fill"])
                    c.border = _thin()
            ws.row_dimensions[evt_row].height = max(18, slot_lines * _LINE_HEIGHT + 4)

        row += _ROWS_PER_WEEK

    # ── Month labels — merged across entire month, alternating colors ──────
    for label in month_order:
        info = month_info[label]
        fill_hex = _MONTH_FILLS[info["color_idx"] % len(_MONTH_FILLS)]
        first_r = info["first_row"]
        last_r = info["last_row"]

        # Fill every cell in the span so borders appear on all rows
        for r in range(first_r, last_r + 1):
            c = ws.cell(row=r, column=_MONTH_COL)
            c.fill = _fill(fill_hex)
            c.border = _thin()

        # Value and formatting on the top cell only (openpyxl merge rule)
        mc = ws.cell(row=first_r, column=_MONTH_COL, value=label)
        mc.font = Font(bold=True, size=11, color="000000")
        mc.fill = _fill(fill_hex)
        mc.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        mc.border = _thin()

        ws.merge_cells(
            start_row=first_r, start_column=_MONTH_COL,
            end_row=last_r, end_column=_MONTH_COL,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
