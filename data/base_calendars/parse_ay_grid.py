#!/usr/bin/env python3
"""
Parse a USMA academic-calendar grid workbook (the "blank calendar" layout
distributed by the Dean's office) into an explicit-days CadetCal JSON.

Grid layout (one row group per week):
  - Seven day groups of three columns each, Sunday..Saturday:
      Sun=B,C,D  Mon=E,F,G  Tue=H,I,J  Wed=K,L,M  Thu=N,O,P  Fri=Q,R,S  Sat=T,U,V
    Within a group: class number ("1-15") in the first column, uniform in the
    second, day-of-month in the third.
  - Rows between one date row and the next hold notes for those same days
    (e.g. "NO CLASSES" / "Labor Day", "TEE", "Reading Day").

Usage:
    python data/base_calendars/parse_ay_grid.py
"""

import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

HERE = Path(__file__).parent

# (class column, date column) per weekday group, 1-based; Mon..Fri only
_WEEKDAY_GROUPS = [(5, 7), (8, 10), (11, 13), (14, 16), (17, 19)]  # E/G, H/J, K/M, N/P, Q/S
_SUNDAY_DATE_COL = 4   # D
_SATURDAY_GROUP = (20, 22)  # T/V

_CLASS_RE = re.compile(r"^([12])-(\d{1,3})$")


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _date_rows(ws) -> list[int]:
    """Rows whose Sunday/Monday date cells hold day-of-month integers."""
    rows = []
    for r in range(1, ws.max_row + 1):
        candidates = [_cell(ws, r, _SUNDAY_DATE_COL)] + [
            _cell(ws, r, dc) for _, dc in _WEEKDAY_GROUPS
        ]
        ints = [c for c in candidates if isinstance(c, (int, float)) and 1 <= int(c) <= 31]
        if len(ints) >= 4:
            rows.append(r)
    return rows


def classify_notes(notes: list[str]) -> str:
    joined = " ".join(notes).lower()
    if "tee" in joined:
        return "tee"
    if "reading day" in joined or "reorgy" in joined or "march back" in joined:
        return "R"
    if "leave" in joined or "break" in joined:
        return "break"
    if "grad" in joined:
        return "grad"
    # "NO CLASSES" plus a holiday name, or an unexplained classless weekday
    return "holiday"


def parse_grid(xlsx_path: Path, ay: str, semester_label: str, first_sunday: date) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    rows = _date_rows(ws)
    blocks = [(r, rows[i + 1] - 1 if i + 1 < len(rows) else ws.max_row)
              for i, r in enumerate(rows)]

    raw_days: dict[date, dict] = {}
    sunday = first_sunday
    for date_row, block_end in blocks:
        # Sanity-check the Sunday day-of-month when present
        sun_val = _cell(ws, date_row, _SUNDAY_DATE_COL)
        if isinstance(sun_val, (int, float)) and int(sun_val) != sunday.day:
            raise ValueError(
                f"Week starting {sunday} expected Sunday day {sunday.day}, "
                f"grid row {date_row} says {int(sun_val)}"
            )
        for offset, (class_col, date_col) in enumerate(_WEEKDAY_GROUPS, start=1):
            d = sunday + timedelta(days=offset)
            dv = _cell(ws, date_row, date_col)
            if isinstance(dv, (int, float)) and int(dv) != d.day:
                raise ValueError(f"Grid row {date_row}: {d} expected day {d.day}, got {int(dv)}")
            class_val = _cell(ws, date_row, class_col)
            notes = [
                str(n) for r in range(date_row + 1, block_end + 1)
                if (n := _cell(ws, r, class_col)) is not None
            ]
            m = _CLASS_RE.match(str(class_val)) if class_val is not None else None
            raw_days[d] = {
                "class": (m.group(1), int(m.group(2))) if m else None,
                "notes": notes,
            }
        # Saturday TEE day (TEEs can run Saturdays)
        sat = sunday + timedelta(days=6)
        sat_notes = [
            str(n) for r in range(date_row + 1, block_end + 1)
            if (n := _cell(ws, r, _SATURDAY_GROUP[0])) is not None
        ]
        if any("tee" in n.lower() for n in sat_notes):
            raw_days[sat] = {"class": None, "notes": ["TEE"]}
        sunday += timedelta(days=7)

    # Determine semester window
    class_days = sorted(d for d, v in raw_days.items() if v["class"])
    tee_days = sorted(
        d for d, v in raw_days.items()
        if not v["class"] and any("tee" in n.lower() for n in v["notes"])
    )
    if not class_days or not tee_days:
        raise ValueError("Could not locate class days or TEE days in grid")
    start, end = class_days[0], class_days[-1]
    tee_start, tee_end = tee_days[0], tee_days[-1]

    days: dict[str, dict] = {}
    d = start
    while d <= max(end, tee_end):
        info = raw_days.get(d)
        if info is None:
            d += timedelta(days=1)
            continue
        if info["class"]:
            track, lesson = info["class"]
            entry = {"day_type": track, "lesson": lesson}
            extra = [n for n in info["notes"] if n.strip()]
            if extra:
                entry["notes"] = extra
            days[d.isoformat()] = entry
        elif d.weekday() < 5 or info["notes"]:
            notes = [n for n in info["notes"] if n.strip()]
            if d.weekday() >= 5 and not notes:
                d += timedelta(days=1)
                continue
            dt = classify_notes(notes)
            # Drop the generic "NO CLASSES" marker, keep the holiday name
            label = [n for n in notes if "no class" not in n.lower()] or notes
            days[d.isoformat()] = {"day_type": dt, "notes": label}
        d += timedelta(days=1)

    return {
        "_comment": f"USMA {ay}. Parsed from the Dean's blank calendar grid. "
                    "Explicit per-date class numbers; rotation is not strictly alternating.",
        "_source": xlsx_path.name,
        "ay": ay,
        "semester": semester_label,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "tee_start": tee_start.isoformat(),
        "tee_end": tee_end.isoformat(),
        "grad_start": None,
        "grad_end": None,
        "days": days,
    }


def main() -> None:
    src = HERE / "sources" / "AY27-1_grid.xlsx"
    data = parse_grid(src, ay="AY27-1", semester_label="Fall 2026",
                      first_sunday=date(2026, 8, 9))
    out = HERE / "AY27-1.json"
    out.write_text(json.dumps(data, indent=2), encoding="utf-8")
    n_academic = sum(1 for v in data["days"].values() if v["day_type"] in ("1", "2"))
    print(f"Wrote {out.name}  (start={data['start_date']}, end={data['end_date']}, "
          f"tee={data['tee_start']}..{data['tee_end']}, "
          f"days={len(data['days'])}, academic={n_academic})")


if __name__ == "__main__":
    main()
