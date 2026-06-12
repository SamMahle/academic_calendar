"""Tests for explicit-days calendars and the structured XLSX extractor.

Synthetic workbooks replicate the real syllabus shapes observed in the wild
(EM384/SE375/EM411/EM381 styles) without shipping anyone's actual syllabus.
"""

import io
from datetime import date, datetime

import openpyxl
import pytest

from core.base_calendar import BaseCalendar
from core.parsers.event_extractor import extract_events
from core.parsers.xlsx_parser import parse_xlsx
from core.parsers.syllabus_xlsx import (
    coerce_date,
    detect_course_code,
    snap_to_semester,
)


def _wb_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Explicit-days calendar
# ---------------------------------------------------------------------------

class TestExplicitCalendar:
    def test_ay27_1_loads_with_40_lessons_each_track(self):
        cal = BaseCalendar("AY27-1")
        assert cal.get_lesson_count(1) == 40
        assert cal.get_lesson_count(2) == 40

    def test_rotation_is_not_strictly_alternating(self):
        # Official AY27-1: Thu 20 Aug = 2-2, Fri 21 Aug = 2-3 (two Day-2s)
        cal = BaseCalendar("AY27-1")
        assert cal.resolve_lesson(2, 2) == date(2026, 8, 20)
        assert cal.resolve_lesson(2, 3) == date(2026, 8, 21)

    def test_ay26_2_explicit_lessons_match_official_card(self):
        # Buff Card: Fri 16 Jan 2026 = 2-5; alternation would get this wrong
        cal = BaseCalendar("AY26-2")
        assert cal.resolve_lesson(2, 5) == date(2026, 1, 16)
        assert cal.resolve_lesson(1, 5) == date(2026, 1, 20)

    def test_ay26_1_unequal_track_lengths(self):
        # Fall 2025 genuinely has 40 Day-1 lessons and 39 Day-2 lessons
        cal = BaseCalendar("AY26-1")
        assert cal.get_lesson_count(1) == 40
        assert cal.get_lesson_count(2) == 39

    def test_saturday_tee_days(self):
        cal1 = BaseCalendar("AY26-1")
        meta = cal1.get_day_meta(date(2025, 12, 20))  # Saturday
        assert meta is not None and meta.day_type == "tee"
        cal2 = BaseCalendar("AY26-2")
        meta = cal2.get_day_meta(date(2026, 5, 16))  # Saturday
        assert meta is not None and meta.day_type == "tee"

    def test_ay27_1_holidays(self):
        cal = BaseCalendar("AY27-1")
        for d in [date(2026, 9, 7), date(2026, 10, 12), date(2026, 11, 26)]:
            meta = cal.get_day_meta(d)
            assert meta is not None and meta.day_type == "holiday", d


# ---------------------------------------------------------------------------
# Date coercion
# ---------------------------------------------------------------------------

class TestDateCoercion:
    @pytest.fixture
    def fall26(self):
        return BaseCalendar("AY27-1")

    def test_compact_day_month(self, fall26):
        assert coerce_date("7SEP", fall26) == date(2026, 9, 7)
        assert coerce_date("20 AUG", fall26) == date(2026, 8, 20)

    def test_weekday_prefixed_text_date(self, fall26):
        assert coerce_date("Mon, 17 Aug", fall26) == date(2026, 8, 17)
        assert coerce_date("Tues, 1 Sept", fall26) == date(2026, 9, 1)
        assert coerce_date("Thurs, 10 Sept", fall26) == date(2026, 9, 10)

    def test_datetime_with_stale_year_is_snapped(self, fall26):
        # Copied template carries 2017; month/day are trusted, year is not
        assert coerce_date(datetime(2017, 9, 14), fall26) == date(2026, 9, 14)

    def test_spring_semester_year_rollover(self):
        spring = BaseCalendar("AY26-2")
        assert coerce_date("9FEB", spring) == date(2026, 2, 9)
        assert coerce_date(datetime(2025, 1, 9), spring) == date(2026, 1, 9)

    def test_unparseable_returns_none(self, fall26):
        assert coerce_date("TBD", fall26) is None
        assert coerce_date(None, fall26) is None

    def test_snap_rejects_far_outside_semester(self, fall26):
        assert snap_to_semester(3, 15, fall26) is None  # March in a fall term


# ---------------------------------------------------------------------------
# Structured extraction fixtures
# ---------------------------------------------------------------------------

def _single_date_col_workbook() -> bytes:
    """EM384 style: one sheet, compact text dates, lesson numbers, blocks."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AT27-1"
    ws.append(["XX301: Test Course Syllabus"])
    ws.append(["Block", "Lsn\n#", "Date\nD1/D2", "Lesson Name",
               "Lesson Assignment", "Graded Event"])
    ws.append(["Block 1", 1, "17AUG", "Intro", "Read Ch 1", None])
    ws.append([None, 2, "19AUG", "Topic A", "Read Ch 2", "Quiz 1"])
    ws.append([None, 3, "24 AUG", "SNOW DAY", None, "Homework Set 1 Due on Canvas"])
    ws.append([None, 4, "26AUG", "WPR 1 Review", "Review L1-3", None])
    ws.append([None, 5, "31AUG", "WPR 1", None, "WPR 1"])
    return _wb_bytes(wb)


def _day1_day2_workbook() -> bytes:
    """SE375 style: Day 1/Day 2 text dates + a graded-events points sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lesson Schedule"
    ws.append(["Lesson", "Day 1", "Day 2", "Topic", "Reading", "Graded Events"])
    ws.append(["Block 1: Basics"])
    ws.append([1, "Mon, 17 Aug", "Tues, 18 Aug", "Course Intro", "Ch 1", None])
    ws.append([2, "Wed, 19 Aug", "Thurs, 20 Aug", "Probability", "Ch 2", "Pre-Test"])
    ws.append([3, "Tues, 8 Sept", "Wed, 9 Sept", "WPR 1", None, "WPR 1"])

    ge = wb.create_sheet("Graded Events")
    ge.append(["Graded Events"])
    ge.append(["Block", "Graded Event", "Type", "Points"])
    ge.append([1, "Pre-Test", "Individual", 50])
    ge.append([1, "WPR 1", "Individual", 150])
    ge.append(["1-4", "TEE", "Individual", 200])
    ge.append([None, None, "Total", 400])
    # Second table with its own header: lesson-indexed quizzes
    ge.append(["Lesson", "Graded Event", "Type", "Points"])
    ge.append([2, "Quiz 1 - Basics", "Quiz", 20])
    return _wb_bytes(wb)


def _per_track_sheets_workbook() -> bytes:
    """EM411 style: per-track sheets, datetime cells with stale years,
    due-date notes."""
    wb = openpyxl.Workbook()
    d1 = wb.active
    d1.title = "Syllabus (Day 1)"
    d1.append(["Block", "LSN", "Date", "Lesson Name", "Notes"])
    d1.append([None, 1, datetime(2026, 8, 17), "Intro", None])
    d1.append([None, 2, datetime(2025, 8, 19), "Planning",
               "Project SG #1 Due by 2359 2 Sep; Template Issued"])
    d1.append([None, None, datetime(2025, 8, 21), None, "Class Drop 1"])
    d1.append([None, 3, datetime(2025, 8, 24), "Quiz #1", "In-Class Quiz, LSN 1-3"])

    d2 = wb.create_sheet("Syllabus (Day 2)")
    d2.append(["Block", "LSN", "Date", "Lesson Name", "Notes"])
    d2.append([None, 1, datetime(2026, 8, 18), "Intro", None])
    d2.append([None, 2, datetime(2025, 8, 20), "Planning",
               "Project SG #1 Due by 2359 2 Sep; Template Issued"])
    d2.append([None, 3, datetime(2025, 8, 25), "Quiz #1", "In-Class Quiz, LSN 1-3"])
    return _wb_bytes(wb)


class TestStructuredExtraction:
    @pytest.fixture
    def fall26(self):
        return BaseCalendar("AY27-1")

    def test_single_date_col(self, fall26):
        doc = parse_xlsx(_single_date_col_workbook())
        events = extract_events(doc, "XX301", 1, fall26)
        by_title = {e.title: e for e in events}
        assert by_title["Quiz 1"].date == date(2026, 8, 19)
        # Snow-day rows still carry their due dates
        assert by_title["Homework Set 1"].date == date(2026, 8, 24)
        assert by_title["WPR 1"].date == date(2026, 8, 31)
        # Lesson topics like "Intro" never become events
        assert all(e.event_type in ("Quiz", "PS", "HW", "WPR") for e in events)

    def test_day1_day2_track_selection(self, fall26):
        doc = parse_xlsx(_day1_day2_workbook())
        track1 = extract_events(doc, "SE999", 1, fall26)
        track2 = extract_events(doc, "SE999", 2, fall26)
        wpr1_t1 = next(e for e in track1 if "WPR 1" in e.title)
        wpr1_t2 = next(e for e in track2 if "WPR 1" in e.title)
        assert wpr1_t1.date == date(2026, 9, 8)
        assert wpr1_t2.date == date(2026, 9, 9)

    def test_points_become_weights(self, fall26):
        doc = parse_xlsx(_day1_day2_workbook())
        events = extract_events(doc, "SE999", 1, fall26)
        wpr1 = next(e for e in events if "WPR 1" in e.title)
        assert wpr1.weight_pct == pytest.approx(37.5)  # 150 / 400
        pre = next(e for e in events if "Pre-Test" in e.title)
        assert pre.weight_pct == pytest.approx(12.5)  # 50 / 400

    def test_second_graded_table_uses_lesson_numbers(self, fall26):
        doc = parse_xlsx(_day1_day2_workbook())
        events = extract_events(doc, "SE999", 1, fall26)
        quiz = next(e for e in events if "Quiz 1" in e.title)
        assert quiz.date == date(2026, 8, 19)  # lesson 2, Day 1

    def test_tee_event_lands_on_tee_week(self, fall26):
        doc = parse_xlsx(_day1_day2_workbook())
        events = extract_events(doc, "SE999", 1, fall26)
        tee = next(e for e in events if e.event_type == "TEE")
        assert tee.date == date(2026, 12, 15)  # first TEE day
        assert tee.confidence < 0.7  # flagged for review

    def test_per_track_sheets_and_stale_years(self, fall26):
        doc = parse_xlsx(_per_track_sheets_workbook())
        t1 = extract_events(doc, "EM999", 1, fall26)
        quiz_t1 = next(e for e in t1 if e.event_type == "Quiz")
        assert quiz_t1.date == date(2026, 8, 24)  # year snapped from 2025
        t2 = extract_events(doc, "EM999", 2, fall26)
        quiz_t2 = next(e for e in t2 if e.event_type == "Quiz")
        assert quiz_t2.date == date(2026, 8, 25)  # Day-2 sheet selected

    def test_due_date_note_overrides_row_date(self, fall26):
        doc = parse_xlsx(_per_track_sheets_workbook())
        events = extract_events(doc, "EM999", 1, fall26)
        sg = next(e for e in events if "SG #1" in e.title)
        assert sg.date == date(2026, 9, 2)

    def test_issued_and_drop_segments_skipped(self, fall26):
        doc = parse_xlsx(_per_track_sheets_workbook())
        events = extract_events(doc, "EM999", 1, fall26)
        assert not any("Issued" in e.title for e in events)
        assert not any("Drop" in e.title for e in events)


class TestCourseCodeDetection:
    def test_from_sheet_title_cell(self):
        doc = parse_xlsx(_single_date_col_workbook())
        assert detect_course_code(doc) == "XX301"

    def test_from_filename_with_underscores(self):
        doc = parse_xlsx(_day1_day2_workbook())
        assert detect_course_code(doc, filename="SE375_271_Syllabus.xlsx") == "SE375"
