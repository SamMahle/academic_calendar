"""End-to-end pipeline tests.

Runs a synthetic DOCX syllabus through the full pipeline and asserts that
extracted events appear in the Excel output.
"""

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from core.base_calendar import BaseCalendar, _build_day_map
from core.models import Course
from core.parsers import ParsedDoc
from core.parsers.event_extractor import extract_events
from core.renderers.excel_renderer import render_excel
from tests.data_helpers import classic_theme

_RAW = {
    "ay": "TEST",
    "semester": "Test",
    "start_date": "2026-01-05",
    "end_date": "2026-04-24",
    "tee_start": "2026-04-27",
    "tee_end": "2026-05-08",
    "grad_start": None,
    "grad_end": None,
    "first_academic_day_type": "1",
    "special_days": {
        "2026-01-19": {"day_type": "holiday", "notes": ["MLK Day"]},
    },
}


@pytest.fixture
def mini_cal():
    return BaseCalendar("TEST", _day_map=_build_day_map(_RAW))


def _make_docx_with_table(rows: list[list[str]]) -> bytes:
    from docx import Document
    doc = Document()
    tbl = doc.add_table(rows=len(rows), cols=len(rows[0]))
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            tbl.cell(i, j).text = val
    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    return buf.read()


class TestFullPipeline:
    def test_docx_table_to_excel(self, mini_cal):
        """Extract WPR from a DOCX table and verify it appears in the Excel output."""
        from core.parsers.docx_parser import parse_docx

        # Build a syllabus with known WPR events
        docx_bytes = _make_docx_with_table([
            ["Event Type", "Lesson", "Weight"],
            ["WPR", "L8", "20%"],
            ["WPR", "L20", "20%"],
            ["TEE", "L40", "30%"],
        ])
        doc = parse_docx(docx_bytes)
        events = extract_events(doc, "MA364", 1, mini_cal)

        assert len(events) >= 2, f"Expected ≥2 events, got {len(events)}"
        wpr_events = [e for e in events if e.event_type == "WPR"]
        assert len(wpr_events) >= 1

        courses = [Course(code="MA364", short_name="DiffEq", track=1, color="4ECDC4")]
        xlsx_bytes = render_excel(events, courses, mini_cal, classic_theme(), "TEST")

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        # Scan all cell values for "MA364"
        all_values = [str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert any("MA364" in v for v in all_values), "Course code MA364 not found in Excel output"

    def test_empty_syllabus_produces_valid_excel(self, mini_cal):
        """An empty document should still produce a structurally valid Excel file."""
        doc = ParsedDoc()
        events = extract_events(doc, "EN101", 1, mini_cal)
        courses = [Course(code="EN101", short_name="English", track=1, color="96CEB4")]

        xlsx_bytes = render_excel(events, courses, mini_cal, classic_theme(), "TEST")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.active.max_row > 2  # at least title + header rows

    def test_ics_pipeline(self, mini_cal):
        """Events extracted from DOCX produce a parseable ICS file."""
        from icalendar import Calendar
        from core.parsers.docx_parser import parse_docx
        from core.renderers.ics_renderer import render_ics

        docx_bytes = _make_docx_with_table([
            ["Event", "Lesson", "Weight"],
            ["WPR", "L5", "20%"],
        ])
        doc = parse_docx(docx_bytes)
        events = extract_events(doc, "MA364", 1, mini_cal)
        courses = [Course(code="MA364", short_name="DiffEq", track=1, color="4ECDC4")]

        ics_bytes = render_ics(events, courses, "TEST")
        cal = Calendar.from_ical(ics_bytes)
        vevents = [c for c in cal.walk() if c.name == "VEVENT"]
        assert len(vevents) == len(events)

    def test_copilot_events_reach_excel(self, mini_cal):
        """Events imported via Copilot paste-back appear in the Excel output."""
        from core.copilot import parse_copilot_table

        md = """
| Event Type | Event Name | Due Date (DD MMM YYYY) | Lesson Reference | Weight (%) |
|---|---|---|---|---|
| WPR | WPR 1 | 04 Feb 2026 | L10 | 20 |
"""
        events = parse_copilot_table(md, "MA364", mini_cal)
        assert len(events) == 1
        assert events[0].date == date(2026, 2, 4)
        assert events[0].source == "copilot"

        courses = [Course(code="MA364", short_name="DiffEq", track=1, color="4ECDC4")]
        xlsx_bytes = render_excel(events, courses, mini_cal, classic_theme(), "TEST")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        all_values = [str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert any("MA364" in v for v in all_values)
